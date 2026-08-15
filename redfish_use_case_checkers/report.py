# Copyright Notice:
# Copyright 2017-2025 Distributed Management Task Force, Inc. All rights reserved.
# License: BSD 3-Clause License. For full text see link: https://github.com/DMTF/Redfish-Use-Case-Checkers/blob/main/LICENSE.md

import html as html_mod
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from redfish_service_validator.html_template import build_html_report
from redfish_use_case_checkers.system_under_test import SystemUnderTest

_RUCC_EXTRA_CSS = r"""

    /* ── Section ── */
    .section { margin-bottom: 1.5rem; }
    .section-header {
      background: #f7f9fc;
      color: #0d1b2a;
      padding: 10px 16px;
      border-radius: 8px 8px 0 0;
      font-weight: 600;
      font-size: 13px;
      letter-spacing: .01em;
      display: flex;
      align-items: center;
      gap: .55rem;
      cursor: pointer;
      user-select: none;
      border-bottom: 1px solid #dde3ec;
    }
    .section-arrow { font-size: .7rem; color: #6c757d; transition: transform .2s; }
    .section-header.collapsed .section-arrow { transform: rotate(-90deg); }
    .section-body {
      border: 1px solid #dde3ec;
      border-top: none;
      border-radius: 0 0 8px 8px;
      overflow: hidden;
      box-shadow: 0 1px 3px rgba(0,0,0,.05);
      margin-bottom: 8px;
    }

    /* ── Test block ── */
    .test-block { border-bottom: 1px solid #dde3ec; }
    .test-block:last-child { border-bottom: none; }
    .test-heading {
      background: #fff;
      padding: 10px 16px;
      cursor: pointer;
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 1rem;
    }
    .test-heading:hover { background: #f8fafc; }
    .test-name   { font-weight: 700; font-size: 13px; color: #0d1b2a; font-family: "Cascadia Code", "Consolas", monospace; }
    .test-desc   { font-size: 11px; color: #6c757d; margin-top: .12rem; font-style: italic; }
    .test-detail { font-size: 11px; color: #6c757d; margin-top: .08rem; font-style: italic; }
    .test-toggle { flex-shrink: 0; font-size: .65rem; color: #6c757d; margin-top: .3rem; transition: transform .2s; }
    .test-body        { display: block; }
    .test-body.hidden { display: none; }

    /* ── Results table ── */
    .test-results-table { width: 100%; border-collapse: collapse; font-size: 12px; }
    .test-results-table thead tr { background: #f0f4f8; }
    .test-results-table th {
      padding: 7px 12px;
      text-align: left;
      font-weight: 700;
      font-size: 11px;
      color: #2c3e50;
      border-bottom: 2px solid #dde3ec;
    }
    .test-results-table td {
      padding: 5px 12px;
      border-bottom: 1px solid #f0f2f5;
      vertical-align: top;
      color: #333;
      background: #fff;
    }
    .test-results-table tr:last-child td { border-bottom: none; }
    .test-results-table tr:hover td { background: #f8fafc; }
    .col-op     { width: 38%; }
    .col-result { width:  9%; white-space: nowrap; }
    .col-msg    { width: 53%; }
    .evidence-list { padding: 8px 12px 10px; background: #f8fafc; border-top: 1px solid #dde3ec; }
    .evidence-call { margin-top: 6px; }
    .evidence-call:first-child { margin-top: 0; }
    .evidence-call-label { align-items: center; background: #f0f4f8; color: #53657d; cursor: pointer; display: flex; font-size: 10px; font-weight: 700; gap: 5px; padding: 4px 6px; }
    .evidence-call-label .uri-arrow { transition: transform .15s; }
    .evidence-call.expanded .uri-arrow { transform: rotate(90deg); }
    .evidence-call-body { display: none; padding: 5px 6px 2px; }
    .evidence-call.expanded .evidence-call-body { display: block; }
    .uri-summary-table { border: 1px solid #d7e0eb; border-collapse: separate; border-spacing: 0; font-size: 11px; width: 100%; }
    .uri-summary-table th { background: #243b53; color: #fff; font-size: 10px; font-weight: 700; letter-spacing: .04em; padding: 7px 10px; text-align: left; text-transform: uppercase; }
    .uri-summary-table th:first-child { border-radius: 4px 0 0 0; }
    .uri-summary-table th:last-child { border-radius: 0 4px 0 0; }
    .uri-summary-table td { border-top: 1px solid #e1e7ef; color: #23395d; padding: 7px 10px; vertical-align: middle; }
    .uri-summary-table .uri-number { font-family: "Cascadia Code", "Consolas", monospace; white-space: nowrap; width: 10%; }
    .uri-summary-table .uri-method { font-weight: 700; width: 10%; }
    .uri-summary-table .uri-status { font-weight: 700; width: 13%; }
    .uri-summary-table .uri-url { overflow-wrap: anywhere; }
    .uri-summary-row { cursor: pointer; }
    .uri-summary-row:hover td { background: #edf3ff; }
    .uri-summary-row.uri-row-open td { background: #eaf2fb; border-top-color: #b8cce2; }
    .uri-row-arrow { color: #58728f; display: inline-block; font-size: 10px; margin-right: 6px; transition: transform .15s; }
    .uri-row-open .uri-row-arrow { transform: rotate(90deg); }
    .uri-method-badge, .uri-status-badge { border-radius: 5px; display: inline-block; font-size: 10px; font-weight: 700; letter-spacing: .04em; padding: 3px 8px; }
    .method-get { background: #e4efff; color: #2456a6; }
    .method-post { background: #dff5e7; color: #17663a; }
    .method-patch { background: #fff0d6; color: #935b00; }
    .method-put { background: #eee6ff; color: #5b35a8; }
    .method-delete { background: #fde3e3; color: #a12626; }
    .method-default { background: #e8edf3; color: #40566f; }
    .status-2xx { background: #dff5e7; color: #17663a; }
    .status-3xx { background: #fff0d6; color: #935b00; }
    .status-4xx { background: #fde3e3; color: #a12626; }
    .status-5xx { background: #f6d5d5; color: #7f1d1d; }
    .status-default { background: #e8edf3; color: #40566f; }
    .uri-detail-row { display: none; }
    .uri-detail-row.open { display: table-row; }
    .uri-detail-row > td { background: #f8fafc; padding: 0 8px 6px; }
    .uri-detail-row .evidence-call { margin: 0; }
    .uri-detail-row .evidence-call-label { display: none; }
    .evidence-buttons { align-items: center; display: flex; flex-wrap: wrap; gap: 7px; }
    .evidence-button {
      align-items: center; background: #fff; border: 1px solid #c7d8f5; border-radius: 5px;
      color: #244a86; cursor: pointer; display: inline-flex; font-size: 11px; font-weight: 700;
      justify-content: center; min-height: 28px; min-width: 92px; padding: 4px 11px;
      transition: background .15s, border-color .15s, box-shadow .15s, color .15s;
    }
    .evidence-button:hover { background: #f4f7fb; border-color: #91acd0; box-shadow: 0 1px 3px rgba(36, 74, 134, .14); }
    .evidence-button.selected { background: #315aa6; border-color: #274a8a; box-shadow: 0 2px 5px rgba(36, 74, 134, .22); color: #fff; }
    .evidence-view-request { border-color: #9bbce8; color: #2456a6; }
    .evidence-view-payload { border-color: #e6c889; color: #805300; }
    .evidence-view-response { border-color: #9bd2ae; color: #17663a; }
    .evidence-view-response-data { border-color: #b8a9e8; color: #51349a; }
    .evidence-button.selected.evidence-view-request { background: #315aa6; color: #fff; }
    .evidence-button.selected.evidence-view-payload { background: #a66a00; color: #fff; }
    .evidence-button.selected.evidence-view-response { background: #21864a; color: #fff; }
    .evidence-button.selected.evidence-view-response-data { background: #6246ad; color: #fff; }
    .evidence-copy { margin-left: auto; min-width: 128px; }
    .evidence-modal-copy { border-color: #91acd0; color: #244a86; }
    .evidence-panel { display: none; }
    .evidence-panel pre {
      max-height: 260px; overflow: auto; margin: 0; padding: 8px;
      background: #132238; color: #e6edf7; border-radius: 4px;
      font: 10px/1.35 "Cascadia Code", "Consolas", monospace; white-space: pre-wrap;
    }
    .evidence pre {
      max-height: 260px; overflow: auto; margin: 4px 0 0; padding: 8px;
      background: #132238; color: #e6edf7; border-radius: 4px;
      font: 10px/1.35 "Cascadia Code", "Consolas", monospace; white-space: pre-wrap;
    }
    .evidence-modal-backdrop {
      align-items: center; background: rgba(13, 27, 42, .48); display: none;
      inset: 0; justify-content: center; padding: 18px; position: fixed; z-index: 1000;
    }
    .evidence-modal-backdrop.open { display: flex; }
    .evidence-modal {
      background: #fff; border: 1px solid #c7d8f5; border-radius: 8px;
      box-shadow: 0 12px 35px rgba(13, 27, 42, .28); max-height: 90vh;
      max-width: 980px; overflow: hidden; width: min(980px, 100%);
    }
    .evidence-modal-header { align-items: center; display: flex; justify-content: space-between; padding: 12px 14px; }
    .evidence-modal-title { color: #20395f; font-size: 13px; font-weight: 700; }
    .evidence-modal-close { background: #fff; border: 1px solid #c7d8f5; border-radius: 5px; color: #244a86; cursor: pointer; font-size: 16px; padding: 3px 9px; }
    .evidence-modal-actions { align-items: center; border-top: 1px solid #dde3ec; display: flex; gap: 7px; padding: 10px 14px; }
    .evidence-modal-views { display: flex; flex-wrap: wrap; gap: 7px; }
    .evidence-modal-copy { margin-left: auto; }
    .evidence-modal-content { margin: 0 14px 14px; max-height: 65vh; overflow: auto; }
    .evidence-modal-content pre { background: #132238; border-radius: 5px; color: #e6edf7; margin: 0; padding: 10px; white-space: pre-wrap; }

    /* ── Badges ── */
    .badge {
      display: inline-block;
      padding: .15rem .55rem;
      border-radius: 9999px;
      font-size: .69rem;
      font-weight: 700;
      letter-spacing: .04em;
      text-transform: uppercase;
    }

    /* ── Toolbar ── */
    .toolbar {
      display: flex;
      align-items: center;
      gap: .6rem;
      margin-bottom: 1rem;
    }
    .toolbar-btn {
      background: #0d6efd;
      color: #fff;
      border: none;
      border-radius: 5px;
      padding: 5px 13px;
      font-size: 11px;
      font-weight: 600;
      cursor: pointer;
      letter-spacing: .02em;
      transition: background .15s;
      box-shadow: 0 2px 5px rgba(13,110,253,.3);
    }
    .toolbar-btn:hover { background: #0b5ed7; }
    .section-toolbar { display: flex; gap: .5rem; padding: 7px 12px; background: #f8fafc; border-bottom: 1px solid #dde3ec; }
    .section-toolbar .toolbar-btn { padding: 4px 10px; font-size: 10px; }

    /* ── Inline result summary chips ── */
    .result-chips { display: flex; gap: .4rem; margin-top: .45rem; flex-wrap: wrap; }
    .rchip {
      display: inline-flex; align-items: center; gap: .3rem;
      padding: .22rem .75rem; border-radius: 9999px;
      font-size: .72rem; font-weight: 700; white-space: nowrap; cursor: default;
    }
    .rchip-pass { background: #c6f6d5; color: #276749; border: 1px solid #38a169; }
    .rchip-warn { background: #fefcbf; color: #744210; border: 1px solid #d69e2e; }
    .rchip-fail { background: #fed7d7; color: #742a2a; border: 1px solid #e53e3e; }
    .rchip-skip { background: #e2e8f0; color: #4a5568; border: 1px solid #a0aec0; }

    /* ── Section count badges ── */
    .section-counts { display: flex; gap: .35rem; margin-left: auto; flex-shrink: 0; align-items: center; }
    .scnt { padding: 2px 8px; border-radius: 20px; font-size: 11px; font-weight: 700; letter-spacing: .01em; }
    .scnt-pass { background: #d4edda; color: #145a32; }
    .scnt-warn { background: #fff3cd; color: #7d6008; }
    .scnt-fail { background: #f8d7da; color: #7b241c; }
    .scnt-skip { background: #f5f7fa; color: #7f8c8d; }

    /* ── Test block status left border ── */
    .test-block.tb-fail { border-left: 4px solid #e53e3e; }
    .test-block.tb-warn { border-left: 4px solid #d69e2e; }
    .test-block.tb-pass { border-left: 4px solid #38a169; }
"""

_RUCC_EXTRA_JS = r"""

  /* Section collapse/expand */
  document.querySelectorAll('.section-header').forEach(function(hdr) {
    hdr.addEventListener('click', function() {
      this.classList.toggle('collapsed');
      var body = this.nextElementSibling;
      body.style.display = (body.style.display === 'none') ? '' : 'none';
    });
  });

  /* Test row collapse/expand */
  document.querySelectorAll('.test-heading').forEach(function(hdr) {
    hdr.addEventListener('click', function() {
      var body = this.nextElementSibling;
      var arrow = this.querySelector('.test-toggle');
      if (body && body.classList.contains('test-body')) {
        var hidden = body.classList.toggle('hidden');
        if (arrow) arrow.style.transform = hidden ? 'rotate(-90deg)' : '';
      }
    });
  });

  /* Expand all */
  function expandAll() {
    document.querySelectorAll('.section-header').forEach(function(hdr) {
      hdr.classList.remove('collapsed');
      hdr.nextElementSibling.style.display = '';
    });
    document.querySelectorAll('.test-body').forEach(function(b) {
      b.classList.remove('hidden');
    });
    document.querySelectorAll('.test-toggle').forEach(function(a) {
      a.style.transform = '';
    });
    document.querySelectorAll('.evidence-call').forEach(function(call) {
      call.classList.add('expanded');
    });
    document.querySelectorAll('.uri-detail-row').forEach(function(row) {
      row.classList.add('open');
    });
    document.querySelectorAll('.section-toolbar .toolbar-btn').forEach(function(button) {
      setToggleLabel(button, true);
    });
  }

  /* Collapse all */
  function collapseAll() {
    document.querySelectorAll('.section-header').forEach(function(hdr) {
      hdr.classList.add('collapsed');
      hdr.nextElementSibling.style.display = 'none';
    });
    document.querySelectorAll('.test-body').forEach(function(b) {
      b.classList.add('hidden');
    });
    document.querySelectorAll('.test-toggle').forEach(function(a) {
      a.style.transform = 'rotate(-90deg)';
    });
    document.querySelectorAll('.evidence-call').forEach(function(call) {
      call.classList.remove('expanded');
    });
    document.querySelectorAll('.uri-detail-row').forEach(function(row) {
      row.classList.remove('open');
    });
    document.querySelectorAll('.section-toolbar .toolbar-btn').forEach(function(button) {
      setToggleLabel(button, false);
    });
  }

  function setToggleLabel(button, expanded) {
    button.dataset.expanded = expanded ? 'true' : 'false';
    button.innerHTML = expanded ? '&#9654;&nbsp; Collapse All' : '&#9660;&nbsp; Expand All';
  }

  function toggleAll(button) {
    var expanded = button.dataset.expanded === 'true';
    if (expanded) {
      collapseAll();
    } else {
      expandAll();
    }
    setToggleLabel(button, !expanded);
  }

  /* URI filter */
  (function() {
    var filter = document.getElementById('uriFilter');
    var clearBtn = document.getElementById('filterClear');
    var countEl = document.getElementById('filterCount');
    if (!filter) return;

    function updateCount() {
      var allTests = document.querySelectorAll('.test-block').length;
      var visibleTests = 0;
      document.querySelectorAll('.test-block').forEach(function(tb) {
        if (tb.style.display !== 'none') visibleTests += 1;
      });
      if (countEl) countEl.innerHTML = '<b>' + visibleTests + '</b> / ' + allTests;
      if (clearBtn) clearBtn.style.display = filter.value.trim() ? 'block' : 'none';
    }

    filter.addEventListener('input', function() {
      var q = (this.value || '').toLowerCase().trim();
      var sections = document.querySelectorAll('.section');

      sections.forEach(function(section) {
        var sectionHdr = section.querySelector('.section-header');
        var testBlocks = section.querySelectorAll('.test-block');
        var sectionMatches = false;

        testBlocks.forEach(function(tb) {
          var heading = tb.querySelector('.test-heading');
          var txt = heading ? heading.textContent.toLowerCase() : '';
          var match = !q || txt.indexOf(q) !== -1;
          tb.style.display = match ? '' : 'none';
          if (match) sectionMatches = true;
        });

        if (sectionHdr && !sectionMatches && q) {
          var secTxt = sectionHdr.textContent.toLowerCase();
          if (secTxt.indexOf(q) !== -1) {
            sectionMatches = true;
            testBlocks.forEach(function(tb) { tb.style.display = ''; });
          }
        }

        section.style.display = sectionMatches || !q ? '' : 'none';
      });

      updateCount();
    });

    updateCount();
  })();

  function clearFilter() {
    document.getElementById('uriFilter').value = '';
    document.getElementById('uriFilter').dispatchEvent(new Event('input'));
  }

  function evidenceModal() {
    var modal = document.getElementById('evidenceModal');
    if (modal) return modal;
    modal = document.createElement('div');
    modal.id = 'evidenceModal';
    modal.className = 'evidence-modal-backdrop';
    modal.innerHTML = '<div class="evidence-modal" role="dialog" aria-modal="true">' +
      '<div class="evidence-modal-header"><span class="evidence-modal-title"></span>' +
      '<button class="evidence-modal-close" onclick="closeEvidence()">&times;</button></div>' +
      '<div class="evidence-modal-actions"><div class="evidence-modal-views"></div>' +
      '<button class="evidence-button evidence-modal-copy" onclick="copyModalEvidence(this)">Copy to Clipboard</button></div>' +
      '<div class="evidence-modal-content"><pre></pre></div></div>';
    modal.addEventListener('click', function(event) {
      if (event.target === modal) closeEvidence();
    });
    document.body.appendChild(modal);
    return modal;
  }

  function showEvidence(panelId, button) {
    var panel = document.getElementById(panelId);
    if (!panel) return;
    var call = panel.closest('.evidence-call');
    if (call) {
      call.querySelectorAll('.evidence-button').forEach(function(item) {
        item.classList.remove('selected');
      });
      call.querySelectorAll('.evidence-button').forEach(function(item) {
        var onclick = item.getAttribute('onclick') || '';
        if (onclick.indexOf("'" + panelId + "'") !== -1) item.classList.add('selected');
      });
    }
    var modal = evidenceModal();
    modal._sourceCall = call;
    var views = modal.querySelector('.evidence-modal-views');
    views.innerHTML = '';
    if (call) {
      call.querySelectorAll('.evidence-button:not(.evidence-copy)').forEach(function(item) {
        var sourceOnclick = item.getAttribute('onclick') || '';
        var match = sourceOnclick.match(/showEvidence\('([^']+)'/);
        if (!match) return;
        var viewButton = document.createElement('button');
        viewButton.className = item.className;
        viewButton.textContent = item.textContent;
        viewButton.setAttribute('onclick', "showEvidence('" + match[1] + "', this)");
        if (match[1] === panelId) viewButton.classList.add('selected');
        views.appendChild(viewButton);
      });
    }
    modal.querySelector('.evidence-modal-title').textContent = button ? button.textContent : 'HTTP Evidence';
    var selectedView = views.querySelector('.selected');
    modal.querySelector('.evidence-modal-title').textContent = selectedView ? selectedView.textContent : (button ? button.textContent : 'HTTP Evidence');
    modal.querySelector('.evidence-modal-content pre').textContent = panel.textContent;
    modal.classList.add('open');
  }

  function closeEvidence() {
    var modal = document.getElementById('evidenceModal');
    if (modal) {
      if (modal._sourceCall) {
        modal._sourceCall.querySelectorAll('.evidence-button').forEach(function(item) {
          item.classList.remove('selected');
        });
      }
      modal.querySelectorAll('.evidence-button').forEach(function(item) {
        item.classList.remove('selected');
      });
      modal._sourceCall = null;
      modal.classList.remove('open');
    }
  }

  function copyModalEvidence(button) {
    var modal = document.getElementById('evidenceModal');
    var text = modal ? modal.querySelector('.evidence-modal-content pre').textContent : '';
    copyText(text, button);
  }

  function copyText(text, button) {
    var copied = function() {
      var original = button.textContent;
      button.textContent = 'Copied';
      setTimeout(function() { button.textContent = original; }, 1200);
    };
    if (navigator.clipboard && navigator.clipboard.writeText) {
      navigator.clipboard.writeText(text).then(copied);
      return;
    }
    var helper = document.createElement('textarea');
    helper.value = text;
    document.body.appendChild(helper);
    helper.select();
    document.execCommand('copy');
    helper.remove();
    copied();
  }

  function copyEvidence(panelId, button) {
    var panel = document.getElementById(panelId);
    var call = button.parentElement.parentElement;
    var activePanel = call.querySelector('.evidence-panel.active');
    if (activePanel) panel = activePanel;
    var text = panel ? panel.textContent : '';
    copyText(text, button);
  }

  function toggleUri(call) {
    call.classList.toggle('expanded');
  }

  function toggleUriRow(rowId, row) {
    var detail = document.getElementById(rowId);
    if (!detail) return;
    detail.classList.toggle('open');
    row.classList.toggle('uri-row-open');
  }

  function expandSection(section) {
    section.querySelectorAll('.test-body').forEach(function(body) {
      body.classList.remove('hidden');
    });
    section.querySelectorAll('.test-toggle').forEach(function(arrow) {
      arrow.style.transform = '';
    });
    section.querySelectorAll('.evidence-call').forEach(function(call) {
      call.classList.add('expanded');
    });
    section.querySelectorAll('.uri-detail-row').forEach(function(row) {
      row.classList.add('open');
    });
  }

  function collapseSection(section) {
    section.querySelectorAll('.test-body').forEach(function(body) {
      body.classList.add('hidden');
    });
    section.querySelectorAll('.test-toggle').forEach(function(arrow) {
      arrow.style.transform = 'rotate(-90deg)';
    });
    section.querySelectorAll('.evidence-call').forEach(function(call) {
      call.classList.remove('expanded');
    });
    section.querySelectorAll('.uri-detail-row').forEach(function(row) {
      row.classList.remove('open');
    });
  }

  function toggleSection(section, button) {
    var expanded = button.dataset.expanded === 'true';
    if (expanded) {
      collapseSection(section);
    } else {
      expandSection(section);
    }
    setToggleLabel(button, !expanded);
  }

  document.addEventListener('keydown', function(event) {
    if (event.key === 'Escape') closeEvidence();
  });
"""

# ── Placeholder kept for any external code that may reference this name ──
html_template = None


def html_report(sut: SystemUnderTest, report_dir, time, tool_version, args=None):
    """
    Creates the HTML report for the system under test

    Args:
        sut: The system under test
        report_dir: The directory for the report
        time: The time the tests finished
        tool_version: The version of the tool

    Returns:
        The path to the HTML report
    """
    file = report_dir / datetime.strftime(time, "RedfishUseCaseCheckersReport_%m_%d_%Y_%H%M%S.html")

    html = ""
    for category_index, test_category in enumerate(sut._results, start=1):
        # Aggregate counts for section header badges
        sec_pass = sum(1 for t in test_category["Tests"] for r in t["Results"] if r["Result"] == "PASS")
        sec_warn = sum(1 for t in test_category["Tests"] for r in t["Results"] if r["Result"] == "WARN")
        sec_fail = sum(1 for t in test_category["Tests"] for r in t["Results"] if r["Result"] == "FAIL")
        sec_skip = sum(1 for t in test_category["Tests"] for r in t["Results"] if not r["Result"])
        sec_cnt = '<div class="section-counts">'
        if sec_pass: sec_cnt += '<span class="scnt scnt-pass">&#10003;&nbsp;{}</span>'.format(sec_pass)
        if sec_warn: sec_cnt += '<span class="scnt scnt-warn">&#9888;&nbsp;{}</span>'.format(sec_warn)
        if sec_fail: sec_cnt += '<span class="scnt scnt-fail">&#10007;&nbsp;{}</span>'.format(sec_fail)
        if sec_skip: sec_cnt += '<span class="scnt scnt-skip">&ndash;&nbsp;{}</span>'.format(sec_skip)
        sec_cnt += '</div>'
        html += '<div class="section">'
        html += '<div class="section-header collapsed"><span class="section-arrow">&#9660;</span>{}. {}{}</div>'.format(
          category_index, html_mod.escape(test_category["Category"]), sec_cnt
        )
        html += '<div class="section-body" style="display: none;">'
        html += (
          '<div class="section-toolbar">'
          '<button class="toolbar-btn" data-expanded="false" onclick="event.stopPropagation(); toggleSection(this.closest(\'.section\'), this)">' \
          '&#9660;&nbsp; Expand All</button>'
          '</div>'
        )

        for test_index, test in enumerate(test_category["Tests"], start=1):
            # Determine worst result for status left-border
            _rvals = [r["Result"] for r in test["Results"]]
            if "FAIL" in _rvals:
                tb_cls = " tb-fail"
            elif "WARN" in _rvals:
                tb_cls = " tb-warn"
            elif "PASS" in _rvals:
                tb_cls = " tb-pass"
            else:
                tb_cls = ""
            html += '<div class="test-block{}">' .format(tb_cls)
            html += '<div class="test-heading">'
            html += '<div class="test-heading-info">'
            html += '<div class="test-name">{}.{} {}</div>'.format(
              category_index, test_index, html_mod.escape(test["Name"])
            )
            html += '<div class="test-desc">{}</div>'.format(html_mod.escape(test["Description"]))
            if test.get("Details"):
                html += '<div class="test-detail">{}</div>'.format(html_mod.escape(test["Details"]))
            # Inline result summary chips — aggregate counts (max 4 chips per test)
            t_pass = sum(1 for _r in test["Results"] if _r["Result"] == "PASS")
            t_warn = sum(1 for _r in test["Results"] if _r["Result"] == "WARN")
            t_fail = sum(1 for _r in test["Results"] if _r["Result"] == "FAIL")
            t_skip = sum(1 for _r in test["Results"] if not _r["Result"])
            html += '<div class="result-chips">'
            if t_pass: html += '<span class="rchip rchip-pass">&#10003; {} Pass</span>'.format(t_pass)
            if t_warn: html += '<span class="rchip rchip-warn">&#9888; {} Warn</span>'.format(t_warn)
            if t_fail: html += '<span class="rchip rchip-fail">&#10007; {} Fail</span>'.format(t_fail)
            if t_skip: html += '<span class="rchip rchip-skip">&ndash; {} N/A</span>'.format(t_skip)
            html += '</div>'
            html += '</div>'  # test-heading-info
            html += '<span class="test-toggle">&#9660;</span>'
            html += '</div>'  # test-heading

            html += '<div class="test-body hidden">'
            html += '<table class="test-results-table">'
            html += (
                '<thead><tr>'
                '<th class="col-op">Operation</th>'
                '<th class="col-result">Result</th>'
                '<th class="col-msg">Message</th>'
                '</tr></thead><tbody>'
            )
            for result in test["Results"]:
                if result["Result"] == "PASS":
                    badge = "badge-pass"
                elif result["Result"] == "WARN":
                    badge = "badge-warn"
                elif result["Result"] == "FAIL":
                    badge = "badge-fail"
                else:
                    badge = "badge-skip"
                operation = result["Operation"] if result["Operation"] else "No testing performed"
                html += '<tr><td>{}</td><td><span class="badge {}">{}</span></td><td>{}</td></tr>'.format(
                    html_mod.escape(operation),
                    badge,
                    html_mod.escape(result["Result"]) if result["Result"] else "N/A",
                    html_mod.escape(result["Message"]),
                )
            html += '</tbody></table>'
            exchanges = []
            exchange_keys = set()
            for result in test["Results"]:
              result_exchanges = result.get("Exchanges") or ([result["Exchange"]] if result.get("Exchange") else [])
              for exchange in result_exchanges:
                    exchange_key = json.dumps(exchange, sort_keys=True, default=str)
                    if exchange_key not in exchange_keys:
                        exchange_keys.add(exchange_key)
                        exchanges.append(exchange)
            if exchanges:
                html += _uri_summary_table(exchanges, category_index, test_index)
            html += '</div>'  # test-body
            html += '</div>'  # test-block

        html += '</div>'  # section-body
        html += '</div>'  # section

    config_rows_html = ""
    if args:
        for key, val in sorted(args.items()):
            if val is None:
                val = ""
            elif isinstance(val, list):
                val = " ".join(str(v) for v in val)
            else:
                val = str(val)
            if key in ("password",):
                val = "********" if val else ""
            config_rows_html += "<tr><td>{}</td><td>{}</td></tr>".format(
                html_mod.escape(key), html_mod.escape(val)
            )

    main_prefix = (
        '<div class="toolbar">'
      '<button class="toolbar-btn" data-expanded="false" onclick="toggleAll(this)">&#9660;&nbsp; Expand All</button>'
        '</div>'
    )

    with open(str(file), "w", encoding="utf-8") as fd:
        fd.write(
            build_html_report(
                page_title="Redfish Use Case Checkers \u2014 Test Report",
                tool_title="Redfish Use Case Checkers",
                filter_placeholder="Filter by Use Case\u2026",
                filter_count_label="test blocks",
                tool_link="https://github.com/DMTF/Redfish-Use-Case-Checkers",
                tool_repo="DMTF/Redfish-Use-Case-Checkers",
                tool_version=tool_version,
                generated_time=time.strftime("%c"),
                sut_host=html_mod.escape(str(sut.rhost)),
                sut_user=html_mod.escape(str(sut.username)),
                sut_password="********",
                sut_product=html_mod.escape(str(sut.product)),
                sut_manufacturer=html_mod.escape(str(sut.manufacturer)),
                sut_model=html_mod.escape(str(sut.model)),
                sut_firmware=html_mod.escape(str(sut.firmware_version)),
                pass_count=sut.pass_count,
                warn_count=sut.warn_count,
                fail_count=sut.fail_count,
                skip_count=sut.skip_count,
                sidebar_extra_html="",
                config_rows_html=config_rows_html,
                extra_css=_RUCC_EXTRA_CSS,
                main_prefix_html=main_prefix,
                main_content_html=html,
                extra_js=_RUCC_EXTRA_JS,
            )
        )
    return file


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _json_text(value):
    if value is None:
        return "null"
    return json.dumps(value, indent=2, sort_keys=True, default=str)


def _uri_summary_table(exchanges, category_index, test_index):
  html = (
    '<table class="uri-summary-table">'
    '<thead><tr><th>S#</th><th>URI</th><th>METHOD</th><th>STATUS CODE</th></tr></thead><tbody>'
  )
  for uri_index, exchange in enumerate(exchanges, start=1):
    response = exchange.get("Response") or {}
    status = response.get("status", "")
    method = str(exchange.get("method", "")).lower()
    method_class = "method-{}".format(method) if method in ("get", "post", "patch", "put", "delete") else "method-default"
    try:
      status_class = "status-{}xx".format(int(status) // 100)
    except (TypeError, ValueError):
      status_class = "status-default"
    if status_class not in ("status-2xx", "status-3xx", "status-4xx", "status-5xx"):
      status_class = "status-default"
    exchange_index = "{}.{}.{}".format(category_index, test_index, uri_index)
    detail_id = "uri-detail-{}".format(exchange_index.replace(".", "-"))
    html += (
            '<tr class="uri-summary-row" onclick="toggleUriRow(\'{}\', this)">'
            '<td class="uri-number"><span class="uri-row-arrow">&#9654;</span> {}.{}.{} </td><td class="uri-url">{}</td>'
          '<td class="uri-method"><span class="uri-method-badge {}">{}</span></td>'
          '<td class="uri-status"><span class="uri-status-badge {}">{}</span></td></tr>'
      '<tr id="{}" class="uri-detail-row"><td colspan="4">{}</td></tr>'
    ).format(
      detail_id,
      category_index,
      test_index,
      uri_index,
      html_mod.escape(str(exchange.get("url", ""))),
      method_class,
      html_mod.escape(str(exchange.get("method", "")).upper()),
      status_class,
      html_mod.escape(str(status)),
      detail_id,
      _exchange_html(exchange, exchange_index, expanded=True),
    )
  return html + '</tbody></table>'


def _exchange_html(exchange, exchange_index, expanded=False):
    request = {key: exchange.get(key) for key in ("headers", "method", "url")}
    request_payload = exchange.get("data")
    response = exchange.get("Response", {})
    response_details = {
      key: value for key, value in response.items() if key not in ("data", "body")
    }
    sections = {
        "request": request,
        "request-payload": request_payload,
      "response": response_details,
        "response-data": response.get("data") if response else exchange.get("Error", ""),
    }
    expanded_class = " expanded" if expanded else ""
    html = '<div class="evidence-call{}">'.format(expanded_class)
    html += '<div class="evidence-call-label" onclick="toggleUri(this.parentElement)">' \
        '<span class="uri-arrow">&#9654;</span>{} -&gt; {} {}</div>'.format(
      exchange_index,
      html_mod.escape(str(exchange.get("method", "")).upper()),
      html_mod.escape(str(exchange.get("url", ""))),
    )
    html += '<div class="evidence-call-body"><div class="evidence-buttons">'
    evidence_views = [("Request", "request")]
    if request_payload not in (None, "", {}, []):
      evidence_views.append(("Request Payload", "request-payload"))
    evidence_views.extend([
      ("Response", "response"),
      ("Response data", "response-data"),
    ])
    for name, key in evidence_views:
        panel_id = "evidence-{}-{}".format(exchange_index, key)
        html += '<button class="evidence-button evidence-view-{}" title="Show {} details" onclick="showEvidence(\'{}\', this)">{}</button>'.format(
          key.replace("request-payload", "payload").replace("response-data", "response-data"),
          name,
            panel_id, name
        )
    html += '</div>'
    for key, value in sections.items():
        panel_id = "evidence-{}-{}".format(exchange_index, key)
        html += '<div id="{}" class="evidence-panel"><pre>{}</pre></div>'.format(
            panel_id, html_mod.escape(_json_text(value))
        )
    return html + '</div></div>'

def _thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def _cell(ws, row, col, value="", bold=False, color=None, fill_hex=None,
          align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Segoe UI", size=10, bold=bold, color=color or "1A1A2E")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _thin_border()
    if fill_hex:
        c.fill = PatternFill("solid", fgColor=fill_hex)
    return c


def xlsx_report(sut: SystemUnderTest, report_dir, time, tool_version):
    """
    Creates the Excel (xlsx) report for the system under test.

    Args:
        sut: The system under test
        report_dir: The directory for the report
        time: The time the tests finished
        tool_version: The version of the tool

    Returns:
        The path to the xlsx report
    """

    file = report_dir / datetime.strftime(time, "RedfishUseCaseCheckersReport_%m_%d_%Y_%H%M%S.xlsx")

    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    HDR_FILL   = "0D1B2A"
    HDR_FONT   = "FFFFFF"
    SUB_FILL   = "1A3A5C"
    ACCENT_FILL= "1565C0"
    META_FILL  = "F0F2F5"
    PASS_FILL  = "D4EDDA"; PASS_FONT  = "145A32"
    WARN_FILL  = "FFF3CD"; WARN_FONT  = "7D6008"
    FAIL_FILL  = "F8D7DA"; FAIL_FONT  = "7B241C"
    SKIP_FILL  = "F5F7FA"; SKIP_FONT  = "7F8C8D"
    CAT_FILL   = "1565C0"
    TEST_FILL  = "F7F9FC"
    SCORE_PASS = "27AE60"
    SCORE_WARN = "D68910"
    SCORE_FAIL = "C0392B"
    SCORE_SKIP = "5D6D7E"

    # Title row  — spans A:D (4 equal columns)
    ws_sum.merge_cells("A1:D1")
    t = ws_sum["A1"]
    t.value = "Redfish Use Case Checkers — Compliance Report"
    t.font = Font(name="Segoe UI", size=14, bold=True, color=HDR_FONT)
    t.fill = PatternFill("solid", fgColor=HDR_FILL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 32

    # Sub-title
    ws_sum.merge_cells("A2:D2")
    s = ws_sum["A2"]
    s.value = "Version: {}    |    Generated: {}".format(tool_version, time.strftime("%c"))
    s.font = Font(name="Segoe UI", size=10, color=HDR_FONT)
    s.fill = PatternFill("solid", fgColor=SUB_FILL)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[2].height = 18

    # Meta info — label in A, value merged B:D
    meta = [
        ("Target System", sut.rhost),
        ("User", sut.username),
        ("Product", sut.product),
        ("Manufacturer", sut.manufacturer),
        ("Model", sut.model),
        ("Firmware", sut.firmware_version),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        _cell(ws_sum, i, 1, label, bold=True, color="6C757D", fill_hex=META_FILL, align="right")
        ws_sum.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        _cell(ws_sum, i, 2, value, fill_hex="FFFFFF")
        # Apply borders to every cell in the merged range so all edges are visible
        for col in range(2, 5):
            ws_sum.cell(row=i, column=col).border = _thin_border()

    # Blank row, then summary counters spanning A:D
    r = 3 + len(meta) + 1
    score_labels = [("✓  PASS", SCORE_PASS), ("⚠  WARN", SCORE_WARN),
                    ("✗  FAIL", SCORE_FAIL), ("–  NOT TESTED", SCORE_SKIP)]
    for col, (label, fill) in enumerate(score_labels, start=1):
        _cell(ws_sum, r, col, label, bold=True, color=HDR_FONT,
              fill_hex=fill, align="center")
        ws_sum.row_dimensions[r].height = 22
    r += 1
    score_data = [
        (sut.pass_count, PASS_FILL, PASS_FONT),
        (sut.warn_count, WARN_FILL, WARN_FONT),
        (sut.fail_count, FAIL_FILL, FAIL_FONT),
        (sut.skip_count, SKIP_FILL, SKIP_FONT),
    ]
    for col, (count, fill, fnt) in enumerate(score_data, start=1):
        _cell(ws_sum, r, col, count, bold=True, color=fnt,
              fill_hex=fill, align="center")
        ws_sum.row_dimensions[r].height = 28

    # Column widths — A (label) narrower, B-D equal data columns
    ws_sum.column_dimensions["A"].width = 20
    for col_letter in ["B", "C", "D"]:
        ws_sum.column_dimensions[col_letter].width = 24

    # Hide grid lines on summary sheet
    ws_sum.sheet_view.showGridLines = False

    # ── Results sheet ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Results")

    # Column definitions: Category | Test Name | Description | Operation | Result | Message | HTTP Evidence
    col_headers = ["Category", "Test Name", "Description", "Operation", "Result", "Message", "HTTP Evidence"]
    col_widths   = [22, 28, 40, 40, 12, 55, 80]

    for col, (hdr, width) in enumerate(zip(col_headers, col_widths), start=1):
        _cell(ws, 1, col, hdr, bold=True, color=HDR_FONT,
              fill_hex=SUB_FILL, align="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Hide grid lines on results sheet
    ws.sheet_view.showGridLines = False

    row = 2
    result_style = {
        "PASS": (PASS_FILL, PASS_FONT),
        "WARN": (WARN_FILL, WARN_FONT),
        "FAIL": (FAIL_FILL, FAIL_FONT),
    }

    for test_category in sut._results:
        category = test_category["Category"]
        for test in test_category["Tests"]:
            test_name = test["Name"]
            description = test["Description"]
            details = test.get("Details", "")
            for result in test["Results"]:
                operation = result["Operation"] or "No testing performed"
                result_val = result["Result"] or "N/A"
                message = result["Message"]
                result_exchanges = result.get("Exchanges") or ([result["Exchange"]] if result.get("Exchange") else [])
                evidence = "\n\n".join(
                  "HTTP call {}\n{}".format(index, _json_text(exchange))
                  for index, exchange in enumerate(result_exchanges, start=1)
                ) or "No HTTP exchange"

                fill, fnt = result_style.get(result_val, (SKIP_FILL, SKIP_FONT))

                _cell(ws, row, 1, category,    fill_hex=CAT_FILL,  color=HDR_FONT, wrap=True)
                _cell(ws, row, 2, test_name,   fill_hex=TEST_FILL, bold=True, wrap=True)
                _cell(ws, row, 3, description, fill_hex=TEST_FILL, wrap=True)
                _cell(ws, row, 4, operation,   wrap=True)
                _cell(ws, row, 5, result_val,  fill_hex=fill, color=fnt,
                      bold=True, align="center")
                _cell(ws, row, 6, message,     wrap=True)
                _cell(ws, row, 7, evidence,     wrap=True)
                ws.row_dimensions[row].height = 30
                row += 1

    wb.save(str(file))
    return file

